from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import datetime
from random import choice
from time import time
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


# Workbook对象属性（工作簿操作）
wb = load_workbook('test.xlsx')  # 已存在的表，创建对象
# print(wb.sheetnames)  # 获取工作簿中的表(列表)
# print(wb.get_sheet_names())
# print(wb.active)  # 获取当前活跃的sheet
# print(wb.worksheets)  # 以列表形式返回所有的sheet
# print(wb.read_only)  # 判断是否以只读模式打开文档
# print(wb.encoding)  # 获取文档的字符集编码
# print(wb.properties)  # 获取文档的元数据，如标题、创建者、创建日期等

# Workbook,cell对象（工作表操作，单元格）
sheet = wb['Sheet1']  # 创建sheet对象
# print(sheet.title)
# print(sheet.max_row, sheet.max_column)
# print(sheet.min_row, sheet.min_column)
# sheet.title = '表1'  # 修改表名
# print(sheet.title)

# res = sheet.cell(row=1, column=1).value  # 根据行列获取单元格的数据
# print(res)
# cell = sheet['B1']
# print(cell)
# print(cell.row, cell.column, cell.value, cell.coordinate)

# 访问单元格所有信息
for row in sheet.rows:
    for cell in row:  # 循环遍历每一个单元格
        print(cell.value, end=',')
    print()
#
# for row in sheet.values:
#     print(*row)

# wss = wb.active  # 激活表
# wss = wb['Sheet2']  # 选定sheet
# wss['B5'] = 'hrhh'  # 写入数据
# wb.save('test.xlsx')


wb1 = Workbook('test1.xlsx')  # 新建表，创建对象
# print(wb1)
# print(type(wb1))
wb1.create_sheet()  # 创建sheet，，默认插在最后,在创建工作表的时候系统自动命名，依次为Sheet, Sheet1, Sheet2 ...
# print(wb1.sheetnames)
wb1.create_sheet(index=0)  # 插在开头
ws = wb1.create_sheet(title='test')
wb1.create_sheet(index=2)
print(wb1.sheetnames)
# wb1.active
# c = ws['A1']  # 读取单元格，如果不存在将在A1新建一个
# print(c)
# ws.cell(1, 1).value = '2222'
# print(ws.cell(1, 1))
# wb1.save('test1.xlsx')
# ws1 = wb1.active
# ws1 = wb1['test']
# ws1['B5'] = 'hrhh'
ws.append(['TIME', 'TITLE', 'A-Z'])  # 第一行插入数据
for i in range(50):
    TIME = datetime.datetime.now().strftime('%H:%M:%S')
    TITLE = str(time())
    A_Z = get_column_letter(choice(range(1, 26)))
    ws.append([TIME, TITLE, A_Z])

# img = Image('111.jpg')
# img.width, img.height = 140, 90
# ws.add_image(img, 'E2')  # 插入图片
wb1.save('test1.xlsx')

# 获取最大行
# row_max = ws.max_row
# # 获取最大列
# con_max = ws.max_column
wb2 = load_workbook('test1.xlsx')
ws2 = wb2['test']
for row in ws2.rows:
    for cell in row:  # 循环遍历每一个单元格
        print(cell.value, end=',')
    print()

